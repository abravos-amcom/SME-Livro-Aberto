from openpyxl import load_workbook

from from_to_handler.models import FonteDeRecursoFromTo


def import_fontes_de_recurso():
    filepath = './from_to_handler/de-paras/DE-PARA Fontes de Recurso.xlsx'
    wb = load_workbook(filepath)
    ws = wb['Plan1']

    row = 2
    row_is_valid = True
    while row_is_valid:
        try:
            code = int(ws['a' + str(row)].value)
        except TypeError:
            row_is_valid = False
            continue
        name = ws['b' + str(row)].value
        group_code = int(ws['c' + str(row)].value)
        group_name = ws['d' + str(row)].value

        fdr = FonteDeRecursoFromTo()
        fdr.code = code
        fdr.name = name
        fdr.group_code = group_code
        fdr.group_name = group_name
        fdr.save()

        row += 1


def run():
    import_fontes_de_recurso()
